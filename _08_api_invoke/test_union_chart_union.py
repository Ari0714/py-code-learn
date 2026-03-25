import glob

import pandas as pd
from pyecharts.charts import Line, Bar, Grid, Page, Scatter
from pyecharts import options as opts
from utils.util import createFilePath
from datetime import datetime, date, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


result_list = []  # 最终输出结果
charts_date = 31

def plot_multi_indicators(df, html_file, stock_name):

    # 计算最近14天的区间起始时间
    recent_start_date = pd.to_datetime(df["date"]).iloc[-1] - pd.Timedelta(days=charts_date)
    recent_date = pd.to_datetime(df["date"]).iloc[-1]

    df["date"] = pd.to_datetime(df["date"])
    df["date_str"] = df["date"].dt.strftime("%Y-%m-%d")

    x = df["date_str"].tolist()

    # --------------------------------------------------------
    # 主图：收盘价
    # --------------------------------------------------------
    chart_close = (
        Line()
        .add_xaxis(x)
        .add_yaxis("Close", df["close"].tolist(), is_smooth=True)
        .set_global_opts(
            title_opts=opts.TitleOpts(
                title="Close Price",  # 标题文字
                pos_top="5%",  # 距图表顶部 5%
                pos_left="left",  # 居中
                title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
            ),
            tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
            legend_opts=opts.LegendOpts(pos_left="left")
        )
    )

    # # --------------------------------------------------------
    # # RSI（两幅：主 RSI + 中心线）
    # # --------------------------------------------------------
    # chart_rsi = (
    #     Line()
    #     .add_xaxis(x)
    #     .add_yaxis("RSI", df["rsi"].tolist(), is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2))
    #     # .add_yaxis("RSI Midline (50)", [50]*len(df), linestyle_opts=opts.LineStyleOpts(width=1, type_="dotted"))
    #         # --- 超卖 30 ---
    #         .add_yaxis(
    #         "",
    #         [30] * len(df),
    #         is_smooth=False,
    #         linestyle_opts=opts.LineStyleOpts(width=1, type_="dotted", color="#888"),
    #         label_opts=opts.LabelOpts(is_show=False),
    #         is_symbol_show=False,
    #     )
    #         # --- 超买 70 ---
    #         .add_yaxis(
    #         "",
    #         [80] * len(df),
    #         is_smooth=False,
    #         linestyle_opts=opts.LineStyleOpts(width=1, type_="dotted", color="#888"),
    #         label_opts=opts.LabelOpts(is_show=False),
    #         is_symbol_show=False,
    #     )
    #     .set_global_opts(
    #         title_opts=opts.TitleOpts(
    #             title="RSI",  # 标题文字
    #             pos_top="25%",  # 距图表顶部 5%
    #             pos_left="left",  # 居中
    #             title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
    #         ),
    #         tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
    #         datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
    #         legend_opts=opts.LegendOpts(pos_left="left")
    #     )
    # )

    # --------------------------------------------------------
    # RSI 背离
    # --------------------------------------------------------
    """
    🔥 仅一个函数：检测无重绘背离 + 绘图 + 输出 HTML
    :param df: 数据必须包含 date, close, rsi
    :param tolerance: 允许误差（默认 0.3%）
    """
    tolerance = 0.003

    top_points = []
    bottom_points = []

    last_price_high_i = 0
    last_price_low_i = 0

    # ========= ★ 无重绘背离算法（逐根计算，永不回看改历史）★ =========
    for i in range(1, len(df)):
        cur_price = df["close"][i]
        cur_rsi = df["rsi"][i]

        if df["date"][i] == recent_date:
            result_list.append({
                "date": 'rsi',
                "stock": stock_name,
                "status": str(round(df["rsi"][i],1))
            })

        # ----（看跌）----
        if cur_price > df["close"][last_price_high_i] * (1 + tolerance) and cur_rsi < df["rsi"][last_price_high_i]:
            top_points.append((df["date"][i], cur_rsi))
            last_price_high_i = i
            # 如果日期在最近14天内
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "rsi↓"
                })
        elif cur_price > df["close"][last_price_high_i]:  # 继续创新高（无背离）
            last_price_high_i = i

        # ---- 底背离（看涨）----
        if cur_price < df["close"][last_price_low_i] * (1 - tolerance) and cur_rsi > df["rsi"][last_price_low_i]:
            bottom_points.append((df["date"][i], cur_rsi))
            last_price_low_i = i
            # 如果日期在最近14天内
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "rsi↑"
                })
        elif cur_price < df["close"][last_price_low_i]:  # 继续创新低（无背离）
            last_price_low_i = i

    # ========= ★ 绘图 ★ =========
    x = df["date"].tolist()
    rsi = df["rsi"].tolist()

    chart_rsi_2 = (
        Line()
            .add_xaxis(x)
            .add_yaxis("RSI", rsi, is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2))
            .add_yaxis("", [30] * len(df), is_symbol_show=False,
                       linestyle_opts=opts.LineStyleOpts(type_="dotted", width=1, color="#777"))
            .add_yaxis("", [80] * len(df), is_symbol_show=False,
                       linestyle_opts=opts.LineStyleOpts(type_="dotted", width=1, color="#777"))
            .set_global_opts(
            title_opts=opts.TitleOpts(
                            title="RSI 无重绘背离",  # 标题文字
                            pos_top="25%",  # 距图表顶部 5%
                            pos_left="left",  # 居中
                            title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
                        ),
            tooltip_opts=opts.TooltipOpts(trigger="axis"),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
            legend_opts=opts.LegendOpts(pos_left="left")
        )
    )

    #   ⭕ 红色倒三角
    if top_points:
        chart = chart_rsi_2.overlap(
            Scatter()
                .add_xaxis([p[0] for p in top_points])
                .add_yaxis(
                "Bearish Divergence",
                [p[1] for p in top_points],
                symbol="triangle", symbol_rotate=180, symbol_size=15,
                itemstyle_opts=opts.ItemStyleOpts(color="red"),
                label_opts=opts.LabelOpts(is_show=False)
            )
        )

    # 底背离 ⭕ 绿色上三角
    if bottom_points:
        chart = chart_rsi_2.overlap(
            Scatter()
                .add_xaxis([p[0] for p in bottom_points])
                .add_yaxis(
                "Bullish Divergence",
                [p[1] for p in bottom_points],
                symbol="triangle", symbol_size=15,
                itemstyle_opts=opts.ItemStyleOpts(color="green"),
                label_opts=opts.LabelOpts(is_show=False)
            )
        )

    # --------------------------------------------------------
    # KD
    # --------------------------------------------------------
    k = df["fast_k"].tolist()
    d = df["fast_d"].tolist()

    buy_x = []
    buy_y = []
    sell_x = []
    sell_y = []

    for i in range(1, len(k)):
        if k[i - 1] < d[i - 1] and k[i] > d[i] and k[i] < 20:  # 买点
            buy_x.append(x[i])
            buy_y.append(k[i])
            # 如果日期在最近14天内
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "kd↑"
                })
        if k[i - 1] > d[i - 1] and k[i] < d[i] and k[i] > 80:  # 卖点
            sell_x.append(x[i])
            sell_y.append(k[i])
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "kd↓"
                })

    chart_kd = (
        Line()
            .add_xaxis(x)
            .add_yaxis("K", k, is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2, color="#1f77b4"))
            .add_yaxis("D", d, is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2, color="#ff7f0e"))
            .add_yaxis("", [20] * len(k), is_symbol_show=False,
                       linestyle_opts=opts.LineStyleOpts(width=1, type_="dotted", color="#888"),
                       label_opts=opts.LabelOpts(is_show=False))
            .add_yaxis("", [80] * len(k), is_symbol_show=False,
                       linestyle_opts=opts.LineStyleOpts(width=1, type_="dotted", color="#888"),
                       label_opts=opts.LabelOpts(is_show=False))
    )

    # --- 添加买卖信号（散点）---
    scatter_buy = (
        Scatter()
            .add_xaxis(buy_x)
            .add_yaxis(
            "Buy",
            buy_y,
            symbol="triangle",
            symbol_size=12,
            itemstyle_opts=opts.ItemStyleOpts(color="#00cc00")
        )
    )

    scatter_sell = (
        Scatter()
            .add_xaxis(sell_x)
            .add_yaxis(
            "Sell",
            sell_y,
            symbol="triangle-down",
            symbol_size=12,
            itemstyle_opts=opts.ItemStyleOpts(color="#ff0000")
        )
    )

    chart_kd = chart_kd.overlap(scatter_buy).overlap(scatter_sell)
    chart_kd.set_global_opts(
        title_opts=opts.TitleOpts(
            title="KD",  # 标题文字
            pos_top="45%",  # 距图表顶部 5%
            pos_left="left",  # 居中
            title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
        ),
        tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
        datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
        legend_opts=opts.LegendOpts(pos_left="left"),
    )

    # --------------------------------------------------------
    # MACD（MACD + SIGNAL + HIST）
    # --------------------------------------------------------
    turn_buy = []  # 看涨（蓝色上三角）
    turn_sell = []  # 看跌（红色下三角）
    hist = df["macd_hist"].tolist()

    for i in range(1, len(hist)):
        # 负 → 正 （看涨）
        if (hist[i] > hist[i - 1]) and (hist[i-1] < hist[i - 2]) and (hist[i-2] < hist[i - 3])\
                 and (hist[i-3] < hist[i - 4]) and (hist[i-4] < hist[i - 5]) and (hist[i-5] < hist[i - 6]):
            turn_buy.append((df["date"][i], df["macd_hist"][i]))
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "macd↑"
                })

        # 正 → 负 （看跌）
        elif (hist[i] < hist[i - 1]) and (hist[i-1] > hist[i - 2]) and (hist[i-2] > hist[i - 3])\
                 and (hist[i-3] > hist[i - 4]) and (hist[i-4] > hist[i - 5]) and (hist[i-5] < hist[i - 6]):
            turn_sell.append((df["date"][i], df["macd_hist"][i]))
            if df["date"][i] >= recent_start_date:
                result_list.append({
                    "date": str(df["date"][i]),
                    "stock": stock_name,
                    "status": "macd↓"
                })

    chart_macd = (
        Line()
        .add_xaxis(x)
        .add_yaxis("MACD", df["macd"].tolist(), is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2, color="red"))
        .add_yaxis("Signal", df["macd_signal"].tolist(), is_smooth=True, linestyle_opts=opts.LineStyleOpts(width=2, color="green"))
        .add_yaxis("Histogram", df["macd_hist"].tolist(), is_smooth=False, linestyle_opts=opts.LineStyleOpts(width=2, color="#B0B0B0"))
        .set_global_opts(
            title_opts=opts.TitleOpts(
                title="MACD",  # 标题文字
                pos_top="65%",  # 距图表顶部 5%
                pos_left="left",  # 居中
                title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
            ),
            # title_opts=opts.TitleOpts(title="MACD"),
            tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
            legend_opts=opts.LegendOpts(pos_left="left")
        )
    )

    # ===== 覆盖转向信号（Scatter 三角形）=====
    if turn_buy:
        chart_macd = chart_macd.overlap(
            Scatter()
                .add_xaxis([p[0] for p in turn_buy])
                .add_yaxis(
                "MACD Turning Up",
                [p[1] for p in turn_buy],
                symbol="triangle",
                symbol_size=14,
                itemstyle_opts=opts.ItemStyleOpts(color="blue"),
                label_opts=opts.LabelOpts(is_show=False),
            )
        )
    if turn_sell:
        chart_macd = chart_macd.overlap(
            Scatter()
                .add_xaxis([p[0] for p in turn_sell])
                .add_yaxis(
                "MACD Turning Down",
                [p[1] for p in turn_sell],
                symbol="triangle",
                symbol_rotate=180,
                symbol_size=14,
                itemstyle_opts=opts.ItemStyleOpts(color="red"),
                label_opts=opts.LabelOpts(is_show=False),
            )
        )

    # --------------------------------------------------------
    # CCI
    # --------------------------------------------------------
    # chart_cci = (
    #     Line()
    #     .add_xaxis(x)
    #     .add_yaxis("CCI", df["cci"].tolist(), is_smooth=True)
    #     .add_yaxis("Zero", [0]*len(df), linestyle_opts=opts.LineStyleOpts(type_="dotted"))
    #     .set_global_opts(
    #         title_opts=opts.TitleOpts(
    #             title="CCI",  # 标题文字
    #             pos_top="85%",  # 距图表顶部 5%
    #             pos_left="left",  # 居中
    #             title_textstyle_opts=opts.TextStyleOpts(font_size=14, color="#333")
    #         ),
    #         tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
    #         datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
    #         legend_opts=opts.LegendOpts(pos_left="left")
    #     )
    # )

    # --------------------------------------------------------
    # 统一排列输出
    # --------------------------------------------------------
    grid = Grid(init_opts=opts.InitOpts(width="1400px", height="2400px",page_title=stock_name))

    grid.add(chart_close, grid_opts=opts.GridOpts(pos_bottom="80%"))
    grid.add(chart_rsi_2, grid_opts=opts.GridOpts(pos_top="22%", pos_bottom="60%"))
    grid.add(chart_kd, grid_opts=opts.GridOpts(pos_top="42%", pos_bottom="40%"))
    grid.add(chart_macd, grid_opts=opts.GridOpts(pos_top="62%", pos_bottom="20%"))
    # grid.add(chart_cci, grid_opts=opts.GridOpts(pos_top="82%"))

    chart_close.set_global_opts(title_opts=opts.TitleOpts(title="Close"))
    chart_rsi_2.set_global_opts(title_opts=opts.TitleOpts(title="RSI"))
    chart_macd.set_global_opts(title_opts=opts.TitleOpts(title="MACD"))
    chart_kd.set_global_opts(title_opts=opts.TitleOpts(title="KD"))
    # chart_cci.set_global_opts(title_opts=opts.TitleOpts(title="CCI"))

    grid.render(html_file)
    print(f"HTML 图像已生成：{html_file}")


def divergence_analysis(result_list, file_date):
    tmp = defaultdict(list)
    # stocks = set()
    stocks = [
        "voo", "qqq", "smh", "","",
        "nvda", "goog", "tsla", "aapl", "meta","","",
        "amd", "tsm", "avgo", "crdo", "mu", "lite","","",
        "iren", "cifr", "nbis", "wulf", "clsk","","",
        "rklb", "asts", "onds","","",
        "be", "te", "oklo", "uuuu","","",
        "hood", "pltr"]
    dates = set()
    # 补充列如果没当日的标记
    dates.add(f'{str(file_date.year)[-2:]}' + f'{file_date.month:02d}' + f'{file_date.day-1:02d}')

    for item in result_list:
        date = item['date'].split(" ")[0][2:].replace('-','')
        stock = item['stock']
        status = item['status']
        tmp[(stock, date)].append(status)
        # stocks.add(stock)
        dates.add(date)

    # 日期正序排列
    dates = sorted(list(dates))

    # 构建矩阵
    matrix = []
    for stock in (list(stocks)):
        row = [stock]
        for date in dates:
            statuses = tmp.get((stock, date), [])
            row.append("\n".join(statuses) if statuses else "")
        matrix.append(row)

    columns = ["stock"] + dates
    df = pd.DataFrame(matrix, columns=columns)

    # 输出Excel
    file_path = f"背离统计-{file_date}.xlsx"
    df.to_excel(file_path, index=False)

    # ====== 填充颜色样式 ======
    # 看涨关键词 → 红框
    bullish_keywords = ["↑"]
    # 看跌关键词 → 绿框
    bearish_keywords = ["↓"]
    # 看跌关键词 & rsi
    bearish_keywords_rsi = ["rsi↓"]
    # 看涨关键词 & rsi
    bullish_keywords_rsi = ["rsi↑"]

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_fill_dark = PatternFill(start_color="33A02C", end_color="33A02C", fill_type="solid")
    red_fill_dark = PatternFill(start_color="D62728", end_color="D62728", fill_type="solid")


    wb = load_workbook(file_path)
    ws = wb.active

    # 状态区：从第2行、第2列开始
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            value = str(cell.value) if cell.value else ""
            if any(k in value for k in bullish_keywords):
                cell.fill = red_fill
            if any(k in value for k in bearish_keywords):
                cell.fill = green_fill
            if any(k in value for k in bearish_keywords_rsi):
                cell.fill = green_fill_dark
            if any(k in value for k in bullish_keywords_rsi):
                cell.fill = red_fill_dark

    wb.save(file_path)
    print(f"Excel 生成成功：{file_path}")



if __name__ == '__main__':

    for stock_name in [
        "voo", "qqq", "smh",
        "nvda", "goog", "tsla", "aapl", "meta",
        "amd", "tsm", "avgo", "crdo", "mu", "lite",
        "iren", "cifr", "nbis", "wulf", "clsk",
        "rklb", "asts", "onds",
        "be", "te", "oklo", "uuuu",
        "hood", "pltr"]:
    # for stock_name in [
    #     "smh"
    #     ]:

        print(f"\n=========={stock_name}============")
        # end_date = datetime.strptime("2025-12-17", "%Y-%m-%d").date()
        # 获取今日日期, 计算去年今日
        end_date = date.today()
        # stock_name = "cifr"
        try:
            df = pd.read_csv(glob.glob(f"output/rsi_union/{end_date.year}/{end_date}/{stock_name}/part-00000-*-c000.csv")[0])

            createFilePath(f"output_html/{end_date}/")
            plot_multi_indicators(df,f"output_html/{end_date}/{stock_name}_{end_date}.html",stock_name)
        except:
            pass

    # ====== 数据整理 ======
    divergence_analysis(result_list, end_date)
